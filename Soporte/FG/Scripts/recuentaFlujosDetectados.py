import sys
import openpyxl
import csv
from openpyxl.styles import Font
import re

def extract_alerts_to_excel(log_file_path):
    # Lista para almacenar las alertas extraídas
    alerts = []
    total_alerts = 0

    # Expresión regular para encontrar líneas con attackid=n, sessionid=n y url
    attackid_pattern = re.compile(r'attackid=\d+')
    sessionid_pattern = re.compile(r'sessionid=(\d+)')
    url_pattern = re.compile(r'url="([^"]+)"')

    # Variables para el control de sessionid y nataques
    sessionid_dict = {}
    nataques = 0
    url_present = False

    # Procesar el archivo de registro
    with open(log_file_path, 'r') as log_file:
        for line in log_file:
            if attackid_pattern.search(line):
                total_alerts += 1

                # Extraer los campos de interés de la línea
                alert = {}
                parts = line.split()
                for part in parts:
                    key_value = part.split("=")
                    if len(key_value) == 2:
                        key = key_value[0]
                        value = key_value[1].strip('"')
                        if key == "attackid":
                            alert[key] = value
                        elif key == "sessionid":
                            sessionid = value
                            alert[key] = sessionid
                            if sessionid not in sessionid_dict:
                                nataques += 1
                                sessionid_dict[sessionid] = nataques
                            else:
                                nataques = sessionid_dict[sessionid]
                        elif key in ["severity", "attack", "service", "srcip", "dstip", "srcport", "dstport", "subtype", "eventtype"]:
                            alert[key] = value
                        elif key == "url":
                            alert[key] = value
                            url_present = True

                if "nataque" not in alert:
                    alert["nataque"] = nataques

                alerts.append(alert)

    # Si no se encontraron alertas, retornar
    if not alerts:
        print("No se encontraron alertas en el archivo de registro.")
        return

    # Contar el número de alertas encontradas
    num_alerts = len(alerts)
    print(f"Se encontraron {num_alerts} alertas únicas en el archivo de registro.")
    print(f"Se encontraron {total_alerts} alertas totales en el archivo de registro.")

    # Crear un libro y hoja de Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Alertas'

    # Escribir encabezados en negrita
    headers = ["attackid", "nataque", "severity", "attack", "service", "srcip", "dstip", "srcport", "dstport", "subtype", "eventtype", "sessionid"]
    if url_present:
        headers.append("url")
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = bold_font

    # Escribir datos
    for row_idx, alert in enumerate(alerts, start=2):
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = alert.get(header, "")

    # Ajustar el tamaño de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener el nombre de la columna
        for cell in col:
            try:  # Necesario para evitar errores en celdas vacías
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar el libro de Excel
    excel_file_path = log_file_path.replace('.log', '.xlsx')
    wb.save(excel_file_path)
    print(f"Las alertas con attackid se han guardado en {excel_file_path}")

    # Convertir Excel a CSV
    csv_file_path = excel_file_path.replace('.xlsx', '.csv')
    with open(csv_file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)  # Escribir encabezados
        for alert in alerts:
            writer.writerow([alert.get(header, "") for header in headers])

    print(f"{excel_file_path} se ha convertido a {csv_file_path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python script.py <ruta_archivo_log>")
        sys.exit(1)

    log_file_path = sys.argv[1]
    extract_alerts_to_excel(log_file_path)
