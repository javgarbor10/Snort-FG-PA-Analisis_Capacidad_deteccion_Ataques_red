import sys
import openpyxl
import csv
import re
from openpyxl.styles import Font

def extract_alerts_to_excel(log_file_path):
    alerts = []
    total_alerts = 0

    # Expresión regular para encontrar "repeated n times"
    repeated_pattern = re.compile(r'(.*)repeated (\d+) times(.*)')
    
    # Inicializamos variables para determinar si los campos están presentes
    url_present = False
    srcport_present = False
    dstport_present = False

    # Procesar el archivo de registro
    with open(log_file_path, 'r') as log_file:
        for line in log_file:
            # Ignorar líneas que no son de nivel "alert"
            if 'level="alert"' not in line:
                continue

            total_alerts += 1
            repeated_value = 1  # Valor por defecto de "repeated"
            repeated_match = repeated_pattern.search(line)

            if repeated_match:
                line = repeated_match.group(1) + repeated_match.group(3)
                repeated_value = int(repeated_match.group(2))

            # Extraer los campos de interés de la línea
            alert = {}
            parts = line.split()
            for part in parts:
                key_value = part.split("=")
                if len(key_value) == 2:
                    key = key_value[0]
                    value = key_value[1].strip('"')
                    if key in ["attackid", "sessionid", "repeated", "severity", "attack", "service", "srcip", "dstip", "subtype", "eventtype", "url", "srcport", "dstport"]:
                        alert[key] = value

            # Añadir el valor "repeated" a la alerta
            alert["repeated"] = repeated_value

            # Comprobamos si existen los campos para definir su presencia
            if "url" in alert:
                url_present = True
            if "srcport" in alert:
                srcport_present = True
            if "dstport" in alert:
                dstport_present = True

            # Añadir la alerta a la lista
            alerts.append(alert)

    # Si no se encontraron alertas, retornar
    if not alerts:
        print("No se encontraron alertas en el archivo de registro.")
        return

    num_unique_alerts = len(alerts)
    print(f"Se encontraron {total_alerts} alertas en el archivo de registro.")
    print(f"Se registrarán {num_unique_alerts} alertas en el archivo de Excel.")

    # Crear un libro y hoja de Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Alertas'

    # Definir los encabezados
    headers = ["attackid", "sessionid", "repeated", "severity", "attack", "service", "srcip", "dstip"]
    if srcport_present:
        headers.append("srcport")
    if dstport_present:
        headers.append("dstport")
    headers.extend(["subtype", "eventtype"])
    if url_present:
        headers.append("url")  # Añadir "url" si está presente en algún log

    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = bold_font

    # Escribir datos en la hoja de Excel
    for row_idx, alert in enumerate(alerts, start=2):
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = alert.get(header, "")

    # Ajustar el tamaño de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener el nombre de la columna
        for cell in col:
            try:  # Necesario para evitar errores en celdas vacías
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar el libro de Excel
    excel_file_path = log_file_path.replace('.log', '.xlsx')
    wb.save(excel_file_path)
    print(f"Las alertas con level='alert' se han guardado en {excel_file_path}")

    # Convertir Excel a CSV
    csv_file_path = excel_file_path.replace('.xlsx', '.csv')
    with open(csv_file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)  # Escribir encabezados
        for alert in alerts:
            writer.writerow([alert.get(header, "") for header in headers])

    print(f"{excel_file_path} se ha convertido a {csv_file_path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python script.py <ruta_archivo_log>")
        sys.exit(1)

    log_file_path = sys.argv[1]
    extract_alerts_to_excel(log_file_path)