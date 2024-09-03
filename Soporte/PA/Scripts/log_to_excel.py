import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys

def extract_info_from_file(file_path):
    """
    Función para extraer la información de interés de un archivo de texto.
    Extrae campos específicos sin filtrar por log_type.
    """
    extracted_info = []
    
    with open(file_path, 'r') as file:
        for line in file:
            info = {}
            info['log_type'] = extract_value(line, 'log_type')
            info['log_subtype'] = extract_value(line, 'log_subtype')
            threat_id = extract_value(line, 'threat_id')
            info['threat_id'] = threat_id
            info['threat_id2'] = extract_number_in_parentheses(threat_id)
            info['severity_number'] = extract_value(line, 'severity_number')
            info['src_ip'] = extract_value(line, 'src_ip')
            info['src_port'] = extract_value(line, 'src_port')
            info['dst_ip'] = extract_value(line, 'dst_ip')
            info['dst_port'] = extract_value(line, 'dst_port')
            info['extra_data'] = extract_value(line, 'extra_data', True)
            repeatcnt = extract_value(line, 'repeatcnt')
            info['repeatcnt'] = int(repeatcnt) if repeatcnt else 0
            extracted_info.append(info)
    
    return extracted_info

def extract_value(line, field, handle_double_quotes=False):
    """
    Función para extraer el valor de un campo específico en una línea.
    """
    if handle_double_quotes:
        pattern = f'{field}=""([^"]*)""'
    else:
        pattern = f'{field}="([^"]*)"'
    
    match = re.search(pattern, line)
    return match.group(1) if match else None

def extract_number_in_parentheses(value):
    """
    Función para extraer el número entre paréntesis en un valor dado.
    """
    match = re.search(r'\((\d+)\)', value)
    return match.group(1) if match else None

def save_to_excel(filename, info, folder_path):
    """
    Función para guardar la información extraída en un archivo Excel.
    """
    df = pd.DataFrame(info)
    
    if df.empty:
        print(f"El archivo {filename} está vacío o no contiene datos válidos. Se omite.")
        return  # Salir de la función si no hay datos que guardar
    
    # Crear el nombre del archivo Excel basado en el nombre del archivo de texto
    excel_filename = f"{os.path.splitext(filename)[0]}.xlsx"
    output_file = os.path.join(folder_path, excel_filename)
    
    # Guardar el DataFrame en un archivo Excel
    df.to_excel(output_file, index=False)

    # Abrir el archivo Excel con openpyxl para agregar la fórmula
    add_sum_formula(output_file, df)

def add_sum_formula(file_path, df):
    """
    Función para agregar una fórmula de suma en las celdas M2, M3, M4 y M5 del archivo Excel.
    M2: Suma total de 'repeatcnt'.
    M3: Suma de 'repeatcnt' solo para log_type="TRAFFIC".
    M4: Suma de 'repeatcnt' solo para log_type="THREAT".
    M5: Suma de 'repeatcnt' para todos los log_type que no sean "TRAFFIC" ni "THREAT".
    """
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Verificar si 'repeatcnt' existe en df
    if 'repeatcnt' in df.columns:
        # Calcular la columna que corresponde a 'repeatcnt'
        repeatcnt_col = get_column_letter(df.columns.get_loc('repeatcnt') + 1)
        logtype_col = get_column_letter(df.columns.get_loc('log_type') + 1)
        
        # Crear la fórmula de suma para la columna 'repeatcnt' en M2
        sum_formula = f"=SUM({repeatcnt_col}2:{repeatcnt_col}{len(df) + 1})"
        ws['M2'] = sum_formula
        
        # Crear la fórmula de suma para 'repeatcnt' donde log_type="TRAFFIC" en M3
        traffic_sum_formula = f"=SUMIFS({repeatcnt_col}2:{repeatcnt_col}{len(df) + 1}, {logtype_col}2:{logtype_col}{len(df) + 1}, \"TRAFFIC\")"
        ws['M3'] = traffic_sum_formula

        # Crear la fórmula de suma para 'repeatcnt' donde log_type="THREAT" en M4
        threat_sum_formula = f"=SUMIFS({repeatcnt_col}2:{repeatcnt_col}{len(df) + 1}, {logtype_col}2:{logtype_col}{len(df) + 1}, \"THREAT\")"
        ws['M4'] = threat_sum_formula

        # Crear la fórmula de suma para 'repeatcnt' donde log_type no es "TRAFFIC" ni "THREAT" en M5
        other_sum_formula = f"=SUMPRODUCT(({logtype_col}2:{logtype_col}{len(df) + 1}<>\"TRAFFIC\")*({logtype_col}2:{logtype_col}{len(df) + 1}<>\"THREAT\"), {repeatcnt_col}2:{repeatcnt_col}{len(df) + 1})"
        ws['M5'] = other_sum_formula
    else:
        print("'repeatcnt' column is missing in the DataFrame. Formulas cannot be added.")
    
    # Guardar el archivo Excel
    wb.save(file_path)

def process_file(file_path):
    """
    Función principal para procesar un único archivo de log.
    """
    extracted_info = extract_info_from_file(file_path)
    filename = os.path.basename(file_path)
    save_to_excel(filename, extracted_info, os.path.dirname(file_path))

def process_directory(directory_path):
    """
    Función para procesar todos los archivos de log en un directorio.
    """
    for filename in os.listdir(directory_path):
        if filename.endswith('.log'):
            file_path = os.path.join(directory_path, filename)
            process_file(file_path)

if __name__ == "__main__":
    directory_path = r'C:\Users\34693\OneDrive\Escritorio\TFG\APUNTES\PALOALTO\LOGS\Red2\DEFINITIVOS\DEFINITIVOS_NUEVOS'
    
    if not os.path.isdir(directory_path):
        print(f"El directorio {directory_path} no existe.")
        sys.exit(1)
    
    process_directory(directory_path)
